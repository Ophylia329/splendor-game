"""
数据加载器 - 从CSV/Excel加载卡牌和贵族数据
"""
import csv
import os
from typing import List, Dict
from models import Card, Noble
import config


def load_cards_from_csv(csv_path: str) -> List[Card]:
    """
    从CSV文件加载卡牌数据

    CSV格式:
    card_id,level,color,cost_white,cost_blue,cost_green,cost_red,cost_black,points

    参数:
        csv_path: CSV文件路径

    返回:
        List[Card]: 卡牌列表
    """
    cards = []

    if not os.path.exists(csv_path):
        print(f"警告: 找不到文件 {csv_path}，使用测试数据")
        return get_mock_cards()

    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                card = Card(
                    card_id=int(row['card_id']),
                    level=int(row['level']),
                    color=row['color'],
                    cost={
                        'white': int(row.get('cost_white', 0)),
                        'blue': int(row.get('cost_blue', 0)),
                        'green': int(row.get('cost_green', 0)),
                        'red': int(row.get('cost_red', 0)),
                        'black': int(row.get('cost_black', 0)),
                    },
                    points=int(row.get('points', 0))
                )
                cards.append(card)
    except Exception as e:
        print(f"加载CSV文件出错: {e}")
        return get_mock_cards()

    return cards


def load_cards_from_excel(excel_path: str) -> List[Card]:
    """
    从Excel文件加载卡牌数据

    参数:
        excel_path: Excel文件路径

    返回:
        List[Card]: 卡牌列表
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

        cards = []
        # 跳过表头，从第2行开始读取
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # 跳过空行
                continue

            card = Card(
                card_id=int(row[0]),
                level=int(row[1]),
                color=row[2],
                cost={
                    'white': int(row[3] or 0),
                    'blue': int(row[4] or 0),
                    'green': int(row[5] or 0),
                    'red': int(row[6] or 0),
                    'black': int(row[7] or 0),
                },
                points=int(row[8] or 0)
            )
            cards.append(card)

        wb.close()
        return cards

    except ImportError:
        print("错误: 未安装openpyxl库，请运行: pip install openpyxl")
        return get_mock_cards()
    except Exception as e:
        print(f"加载Excel文件出错: {e}")
        return get_mock_cards()


def load_nobles_from_csv(csv_path: str) -> List[Noble]:
    """
    从CSV文件加载贵族数据

    CSV格式:
    noble_id,name,req_white,req_blue,req_green,req_red,req_black

    参数:
        csv_path: CSV文件路径

    返回:
        List[Noble]: 贵族列表
    """
    nobles = []

    if not os.path.exists(csv_path):
        print(f"警告: 找不到文件 {csv_path}，使用测试数据")
        return get_mock_nobles()

    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                noble = Noble(
                    noble_id=int(row['noble_id']),
                    name=row.get('name', f"贵族{row['noble_id']}"),
                    requirements={
                        'white': int(row.get('req_white', 0)),
                        'blue': int(row.get('req_blue', 0)),
                        'green': int(row.get('req_green', 0)),
                        'red': int(row.get('req_red', 0)),
                        'black': int(row.get('req_black', 0)),
                    }
                )
                nobles.append(noble)
    except Exception as e:
        print(f"加载CSV文件出错: {e}")
        return get_mock_nobles()

    return nobles


def load_nobles_from_excel(excel_path: str, sheet_name: str = "Nobles") -> List[Noble]:
    """
    从Excel文件加载贵族数据

    参数:
        excel_path: Excel文件路径
        sheet_name: 工作表名称

    返回:
        List[Noble]: 贵族列表
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        nobles = []
        # 跳过表头，从第2行开始读取
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # 跳过空行
                continue

            noble = Noble(
                noble_id=int(row[0]),
                name=row[1] if row[1] else f"贵族{row[0]}",
                requirements={
                    'white': int(row[2] or 0),
                    'blue': int(row[3] or 0),
                    'green': int(row[4] or 0),
                    'red': int(row[5] or 0),
                    'black': int(row[6] or 0),
                }
            )
            nobles.append(noble)

        wb.close()
        return nobles

    except ImportError:
        print("错误: 未安装openpyxl库，请运行: pip install openpyxl")
        return get_mock_nobles()
    except Exception as e:
        print(f"加载Excel文件出错: {e}")
        return get_mock_nobles()


def get_mock_cards() -> List[Card]:
    """
    生成测试用的卡牌数据
    包含少量各等级的卡牌用于测试

    返回:
        List[Card]: 测试卡牌列表
    """
    mock_cards = [
        # 等级1 - 10张测试卡（恢复正常声望值0-1分）
        Card(1, 1, 'white', {'blue': 3}, 0),
        Card(2, 1, 'white', {'green': 2, 'red': 1}, 0),
        Card(3, 1, 'blue', {'white': 2, 'black': 1}, 0),
        Card(4, 1, 'blue', {'red': 3}, 0),
        Card(5, 1, 'green', {'white': 1, 'blue': 2, 'red': 1, 'black': 1}, 0),
        Card(6, 1, 'green', {'black': 2, 'white': 1}, 0),
        Card(7, 1, 'red', {'white': 3}, 1),
        Card(8, 1, 'red', {'blue': 2, 'green': 1}, 0),
        Card(9, 1, 'black', {'green': 3}, 0),
        Card(10, 1, 'black', {'blue': 1, 'red': 2}, 0),

        # 等级2 - 8张测试卡
        Card(11, 2, 'white', {'green': 3, 'red': 3}, 1),
        Card(12, 2, 'white', {'blue': 2, 'black': 3}, 2),
        Card(13, 2, 'blue', {'white': 3, 'green': 3}, 1),
        Card(14, 2, 'blue', {'red': 2, 'black': 3}, 2),
        Card(15, 2, 'green', {'white': 3, 'blue': 2}, 1),
        Card(16, 2, 'green', {'red': 3, 'black': 2}, 2),
        Card(17, 2, 'red', {'white': 2, 'blue': 3}, 1),
        Card(18, 2, 'black', {'green': 2, 'red': 3}, 2),

        # 等级3 - 6张测试卡
        Card(19, 3, 'white', {'black': 7}, 4),
        Card(20, 3, 'white', {'green': 3, 'red': 3, 'black': 3}, 3),
        Card(21, 3, 'blue', {'white': 7}, 4),
        Card(22, 3, 'green', {'blue': 5, 'red': 3}, 3),
        Card(23, 3, 'red', {'white': 3, 'blue': 3, 'green': 3}, 3),
        Card(24, 3, 'black', {'red': 7}, 5),
    ]

    return mock_cards


def get_mock_nobles() -> List[Noble]:
    """
    生成测试用的贵族数据

    返回:
        List[Noble]: 测试贵族列表
    """
    mock_nobles = [
        Noble(1, "凯瑟琳女王", {'white': 3, 'blue': 3, 'green': 3}, 3),
        Noble(2, "查尔斯国王", {'white': 4, 'red': 4}, 3),
        Noble(3, "伊丽莎白公主", {'blue': 3, 'green': 3, 'red': 3}, 3),
        Noble(4, "弗朗西斯骑士", {'green': 4, 'black': 4}, 3),
        Noble(5, "安妮女爵", {'white': 3, 'blue': 3, 'black': 3}, 3),
        Noble(6, "亨利公爵", {'red': 4, 'black': 4}, 3),
        Noble(7, "玛丽王后", {'white': 3, 'red': 3, 'black': 3}, 3),
        Noble(8, "威廉伯爵", {'blue': 4, 'green': 4}, 3),
        Noble(9, "乔治主教", {'white': 4, 'blue': 4}, 3),
        Noble(10, "维多利亚夫人", {'green': 4, 'red': 4}, 3),
    ]

    return mock_nobles


def save_template_csv():
    """
    生成CSV模板文件，供用户参考

    在data目录下生成cards_template.csv和nobles_template.csv
    """
    # 确保data目录存在
    os.makedirs(config.DATA_DIR, exist_ok=True)

    # 卡牌模板
    cards_template_path = os.path.join(config.DATA_DIR, "cards_template.csv")
    with open(cards_template_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['card_id', 'level', 'color', 'cost_white', 'cost_blue',
                        'cost_green', 'cost_red', 'cost_black', 'points'])
        writer.writerow([1, 1, 'white', 0, 3, 0, 0, 0, 0])
        writer.writerow([2, 1, 'blue', 2, 0, 0, 1, 0, 0])
        writer.writerow([3, 2, 'green', 0, 3, 3, 0, 0, 1])

    # 贵族模板
    nobles_template_path = os.path.join(config.DATA_DIR, "nobles_template.csv")
    with open(nobles_template_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['noble_id', 'name', 'req_white', 'req_blue',
                        'req_green', 'req_red', 'req_black'])
        writer.writerow([1, '贵族1', 3, 3, 3, 0, 0])
        writer.writerow([2, '贵族2', 4, 4, 0, 0, 0])

    print(f"✅ CSV模板已生成:")
    print(f"  - {cards_template_path}")
    print(f"  - {nobles_template_path}")
